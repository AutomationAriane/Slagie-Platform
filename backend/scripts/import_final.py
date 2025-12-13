#!/usr/bin/env python
import os
import sys
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
from openpyxl_image_loader import SheetImageLoader
from PIL import Image

CURRENT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CURRENT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import models
from database import SessionLocal, engine


EXCEL_FILE = PROJECT_ROOT / "data" / "TheorieToppers examen vragen (3).xlsx"
IMAGES_DIR = PROJECT_ROOT / "static" / "images"


def reset_exams(db) -> Tuple[int, int, int]:
    db.query(models.AnswerOption).delete()
    db.query(models.Question).delete()
    db.query(models.SubTopic).delete()
    db.query(models.Topic).delete()
    db.commit()

    exams = [
        models.Topic(name="Examen 1", description="CBR Theorie Examen 1"),
        models.Topic(name="Examen 2", description="CBR Theorie Examen 2"),
        models.Topic(name="Examen 3", description="CBR Theorie Examen 3"),
    ]
    db.add_all(exams)
    db.flush()
    db.commit()
    return (exams[0].id, exams[1].id, exams[2].id)


def parse_index(raw) -> Optional[int]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if "/" in s:
        s = s.split("/")[0]
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def norm_text(x) -> str:
    return " ".join(str(x).split()).strip().lower()


def detect_correct_letter(q_index: int, answer_cell, opt_a, opt_b, opt_c, opt_d) -> Optional[str]:
    target = str(answer_cell).strip() if answer_cell is not None else ""
    cand_map = {
        "A": str(opt_a).strip() if opt_a is not None else None,
        "B": str(opt_b).strip() if opt_b is not None else None,
        "C": str(opt_c).strip() if opt_c is not None else None,
        "D": str(opt_d).strip() if opt_d is not None else None,
    }

    # Exact text match (after strip), A→D
    if target:
        for letter, text in cand_map.items():
            if text is not None and target == text:
                print(f"Vraag {q_index}: Match gevonden op Optie {letter}")
                return letter

    # Fallback: interpret target as letter A-D
    upper = target.upper()
    if upper in {"A", "B", "C", "D"}:
        print(f"Vraag {q_index}: Fallback letter '{upper}' gebruikt")
        return upper

    print(f"⚠️  Vraag {q_index}: Geen match gevonden voor antwoord '{target}'")
    return None


def get_or_create_subtopic(db, exam_id: int, subtopic_name: Optional[str]) -> models.SubTopic:
    name = (subtopic_name or "Algemeen").strip()
    st = (
        db.query(models.SubTopic)
        .filter(models.SubTopic.topic_id == exam_id, models.SubTopic.name == name)
        .first()
    )
    if st:
        return st
    st = models.SubTopic(topic_id=exam_id, name=name, description=None)
    db.add(st)
    db.flush()
    return st


def save_image_from_cell(image_loader: SheetImageLoader, row_idx: int, question_id: int) -> Optional[str]:
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    cell = f"K{row_idx}"
    try:
        img = image_loader.get(cell)
    except Exception:
        img = None
    if img is None:
        return None
    # JPEG does not support alpha; always ensure RGB
    if img.mode != "RGB":
        img = img.convert("RGB")
    img.thumbnail((800, 600))
    filename = f"vraag_{question_id}.jpg"
    target = IMAGES_DIR / filename
    img.save(target, format="JPEG", quality=88)
    return filename


def main():
    models.Base.metadata.create_all(bind=engine)
    if not EXCEL_FILE.exists():
        print(f"❌ Excel bronbestand ontbreekt: {EXCEL_FILE}")
        return

    db = SessionLocal()
    try:
        exam_ids = reset_exams(db)
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        image_loader = SheetImageLoader(ws)

        exam_counter = 0
        imported = 0
        with_images = 0

        for row in range(2, ws.max_row + 1):
            idx = parse_index(ws.cell(row, 1).value)
            text = ws.cell(row, 2).value
            theme = ws.cell(row, 9).value  # I: CBR Thema
            subtopic_name = ws.cell(row, 10).value  # J: Onderwerp
            answer_cell = ws.cell(row, 3).value  # C: Antwoord (tekst of letter)
            opt_a = ws.cell(row, 4).value
            opt_b = ws.cell(row, 5).value
            opt_c = ws.cell(row, 6).value
            opt_d = ws.cell(row, 7).value

            if not text or idx is None:
                continue

            exam_id = exam_ids[exam_counter % 3]
            exam_counter += 1

            st = get_or_create_subtopic(db, exam_id, subtopic_name)

            q = models.Question(
                question_number=idx,
                subtopic_id=st.id,
                text=str(text),
                question_type="multiple_choice",
                image_url=None,
                theme=str(theme) if theme else None,
            )
            db.add(q)
            db.flush()

            img_filename = save_image_from_cell(image_loader, row, q.id)
            if img_filename:
                q.image_url = img_filename
                with_images += 1

            corr = detect_correct_letter(idx, answer_cell, opt_a, opt_b, opt_c, opt_d)
            options = [
                ("A", opt_a, corr == "A"),
                ("B", opt_b, corr == "B"),
                ("C", opt_c, corr == "C"),
                ("D", opt_d, corr == "D"),
            ]
            for letter, t, is_corr in options:
                if t and str(t).strip():
                    db.add(models.AnswerOption(
                        question_id=q.id,
                        option_letter=letter,
                        text=str(t),
                        is_correct=bool(is_corr),
                    ))

            imported += 1
            if imported % 25 == 0:
                print(f"  ✓ {imported} vragen verwerkt...")

        db.commit()
        print("\n============================================================")
        print("✅ Import afgerond")
        print(f"  - Vragen: {imported}")
        print(f"  - Met afbeeldingen: {with_images}")
        print("============================================================")
    finally:
        db.close()


if __name__ == "__main__":
    main()
