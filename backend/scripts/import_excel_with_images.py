#!/usr/bin/env python
"""
Import script: Vult database vanuit Excel
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from database import SessionLocal, engine
import models

def main():
    print("=" * 60)
    print("🚀 Slagie Import Script")
    print("=" * 60)
    
    # Create tables
    models.Base.metadata.create_all(bind=engine)
    print("✅ Database schema created")
    
    # Create new session
    db = SessionLocal()
    
    try:
        # Delete all old data
        print("🗑️  Clearing old data...")
        db.query(models.AnswerOption).delete()
        db.query(models.Question).delete()
        db.query(models.SubTopic).delete()
        db.query(models.Topic).delete()
        db.commit()
        
        # Create 3 exams
        print("📝 Creating 3 exams...")
        exam1 = models.Topic(name="Examen 1", description="CBR Theorie Examen 1 - 50 vragen")
        exam2 = models.Topic(name="Examen 2", description="CBR Theorie Examen 2 - 50 vragen")
        exam3 = models.Topic(name="Examen 3", description="CBR Theorie Examen 3 - 50 vragen")
        db.add(exam1)
        db.add(exam2)
        db.add(exam3)
        db.flush()  # Get IDs without commit
        print(f"✅ Created exams: ID={exam1.id}, ID={exam2.id}, ID={exam3.id}")
        
        # Store exam IDs for later use
        exam_ids = {
            (11, 60): exam1.id,      # Questions 11-60 → Exam 1
            (110, 159): exam2.id,    # Questions 110-159 → Exam 2
            (210, 259): exam3.id,    # Questions 210-259 → Exam 3
        }
        db.commit()
        
        # Import Excel data
        print("\n📚 Importing questions from Excel...")
        excel_file = "/app/data/TheorieToppers examen vragen (3).xlsx"
        
        if not os.path.exists(excel_file):
            print(f"❌ File not found: {excel_file}")
            return
        
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        questions_imported = 0
        answers_imported = 0
        
        for row_idx in range(2, ws.max_row + 1):
            # Parse row
            question_num_raw = ws[f'A{row_idx}'].value
            question_text = ws[f'B{row_idx}'].value
            correct_answer_text = ws[f'C{row_idx}'].value
            answer_a = ws[f'D{row_idx}'].value
            answer_b = ws[f'E{row_idx}'].value
            answer_c = ws[f'F{row_idx}'].value
            answer_d = ws[f'G{row_idx}'].value
            theme = ws[f'I{row_idx}'].value or "Algemeen"
            
            if not question_text or not question_num_raw:
                continue
            
            # Parse question number (11/50 → 11)
            try:
                q_num = int(str(question_num_raw).split('/')[0])
            except:
                continue
            
            # Assign to exam
            if 11 <= q_num <= 60:
                exam_id = exam_ids[(11, 60)]
            elif 110 <= q_num <= 159:
                exam_id = exam_ids[(110, 159)]
            elif 210 <= q_num <= 259:
                exam_id = exam_ids[(210, 259)]
            else:
                continue
            
            # Determine correct answer
            correct_letter = "A"
            if answer_a and correct_answer_text and str(answer_a).strip() == str(correct_answer_text).strip():
                correct_letter = "A"
            elif answer_b and correct_answer_text and str(answer_b).strip() == str(correct_answer_text).strip():
                correct_letter = "B"
            elif answer_c and correct_answer_text and str(answer_c).strip() == str(correct_answer_text).strip():
                correct_letter = "C"
            elif answer_d and correct_answer_text and str(answer_d).strip() == str(correct_answer_text).strip():
                correct_letter = "D"
            
            # Create SubTopic
            subtopic = models.SubTopic(topic_id=exam_id, name=theme, description=f"Vraag {q_num}")
            db.add(subtopic)
            db.flush()
            
            # Create Question
            question = models.Question(
                question_number=q_num,
                subtopic_id=subtopic.id,
                text=question_text,
                question_type="multiple_choice",
                image_path=None,
                theme=theme
            )
            db.add(question)
            db.flush()
            
            # Create AnswerOptions
            answers = [
                ("A", answer_a, correct_letter == "A"),
                ("B", answer_b, correct_letter == "B"),
                ("C", answer_c, correct_letter == "C"),
                ("D", answer_d, correct_letter == "D"),
            ]
            
            for letter, text, is_correct in answers:
                if text:
                    option = models.AnswerOption(
                        question_id=question.id,
                        option_letter=letter,
                        text=str(text),
                        is_correct=is_correct
                    )
                    db.add(option)
                    answers_imported += 1
            
            questions_imported += 1
            if questions_imported % 20 == 0:
                print(f"  ✓ {questions_imported} questions...")
        
        db.commit()
        print(f"\n✅ {questions_imported} questions imported")
        print(f"✅ {answers_imported} answers imported")
        
        # Verify
        print("\n🔍 Verification:")
        for exam in db.query(models.Topic).all():
            count = db.query(models.Question).join(
                models.SubTopic, models.Question.subtopic_id == models.SubTopic.id
            ).filter(models.SubTopic.topic_id == exam.id).count()
            print(f"  - {exam.name}: {count} questions")
        
        print("\n" + "=" * 60)
        print("✅ Import completed!")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    main()
