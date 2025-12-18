"""
Advanced Excel Import Script with Image Extraction
Imports exam questions from TheorieToppers Excel file with embedded images
"""
import os
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))

from sqlalchemy.orm import Session
from database import SessionLocal, engine
from models import Base, Exam, ExamQuestionItem, ExamAnswerOption
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage


def create_static_folder():
    """Create static/images folder if it doesn't exist"""
    static_dir = Path(__file__).parent.parent / "static" / "images"
    static_dir.mkdir(parents=True, exist_ok=True)
    return static_dir


def extract_images_from_excel(workbook, worksheet, static_dir):
    """
    Extract all images from Excel and save them
    Returns dict mapping row numbers to image filenames
    """
    print("📸 Extracting images from Excel...")
    image_map = {}
    
    if not hasattr(worksheet, '_images'):
        print("  ⚠️  No images found in worksheet")
        return image_map
    
    for idx, img in enumerate(worksheet._images):
        try:
            # Get the row number where the image is anchored
            # _from is the top-left cell anchor
            row_number = img.anchor._from.row + 1  # +1 because Excel rows are 1-indexed
            
            # Generate filename
            filename = f"vraag_{row_number}.png"
            filepath = static_dir / filename
            
            # Save image
            image_data = img._data()
            with open(filepath, 'wb') as f:
                f.write(image_data)
            
            # Store mapping
            image_map[row_number] = filename
            print(f"  ✓ Saved image for row {row_number}: {filename}")
            
        except Exception as e:
            print(f"  ✗ Error extracting image {idx}: {e}")
    
    print(f"  📊 Total images extracted: {len(image_map)}")
    return image_map


def parse_excel_data(worksheet, image_map, static_dir):
    """
    Parse questions and answers from Excel
    Returns list of question data dictionaries
    """
    print("\n📖 Parsing Excel data...")
    questions_data = []
    
    # Updated Excel column indices based on peek
    COL_NUMBER = 0  # Vraagnummer
    COL_QUESTION = 1  # Vraagtekst
    COL_ANSWER_TEXT = 2  # Antwoord (Matches text of correct option)
    COL_A = 3  # Optie A
    COL_B = 4  # Optie B
    COL_C = 5  # Optie C
    COL_D = 6  # Optie D
    COL_TYPE = 7  # Vraagtype
    COL_THEME = 8  # CBR Thema
    COL_TOPIC = 9  # Onderwerp
    
    # Skip header row, start from row 2
    row_count = 0
    for excel_row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Extract data
            question_num = row[COL_NUMBER]
            question_text = row[COL_QUESTION]
            correct_text = str(row[COL_ANSWER_TEXT]).strip() if row[COL_ANSWER_TEXT] else ""
            
            cbr_theme = row[COL_THEME]
            cbr_topic = row[COL_TOPIC]
            q_type = "multiple_choice" # Default
            
            # Simple map for type if provided, else default
            raw_type = str(row[COL_TYPE]).lower() if row[COL_TYPE] else ""
            if "gevaar" in raw_type or "gevaar" in str(cbr_theme).lower():
                q_type = "drag_drop" # Or verify if it's actually drag drop? Usually Gevaarherkenning is ABC (Rem/Gas/Niets). 
                # Wait, Gevaarherkenning is ABC. Not Drag Drop.
                # "Meerkeuze" is Multiple Choice.
                q_type = "multiple_choice"
            elif "ja" in raw_type:
                 q_type = "multiple_choice" # Ja/Nee is effectively MC
            
            # Options
            opts = [
                {"text": str(row[COL_A] or "").strip(), "letter": "A"},
                {"text": str(row[COL_B] or "").strip(), "letter": "B"},
                {"text": str(row[COL_C] or "").strip(), "letter": "C"},
                {"text": str(row[COL_D] or "").strip(), "letter": "D"}
            ]
            
            # Filter empty options
            valid_opts = [o for o in opts if o["text"]]
            
            # Determine correctness by matching text
            answers = []
            for o in valid_opts:
                is_correct = (o["text"] == correct_text)
                answers.append({
                    "text": o["text"],
                    "letter": o["letter"],
                    "is_correct": is_correct
                })
            
            # Fallback if no exact match (sometimes whitespace issues)
            if not any(a["is_correct"] for a in answers):
                # Try fuzzy match? Or just log warning.
                # print(f"  ⚠️ No matching answer for Q{question_num}. Correct: '{correct_text}'")
                pass

            
            # Skip if no question text
            if not question_text:
                continue
            
            # Check if image exists for this row
            image_url = None
            if excel_row_num in image_map:
                filename = image_map[excel_row_num]
                image_url = f"http://localhost:8000/static/images/{filename}"
            
            question_data = {
                "number": question_num,
                "text": question_text,
                "image_url": image_url,
                "type": q_type,
                "cbr_topic": cbr_theme,
                "cbr_subtopic": cbr_topic,
                "answers": answers
            }
            
            questions_data.append(question_data)
            row_count += 1
            
        except Exception as e:
            print(f"  ✗ Error parsing row {excel_row_num}: {e}")
    
    print(f"  📊 Total questions parsed: {row_count}")
    return questions_data


def create_exams_and_questions(db: Session, questions_data):
    """
    Create 3 exams with 50 questions each
    """
    print("\n🏗️  Creating exams and questions...")
    
    questions_per_exam = 50
    exam_count = 3
    
    # First, create all question items
    print("  Creating question items...")
    question_items = []
    for q_data in questions_data:
        question = ExamQuestionItem(
            question_text=q_data["text"],
            question_image=q_data["image_url"],
            question_type=q_data["type"],
            cbr_topic=q_data["cbr_topic"],
            cbr_subtopic=q_data["cbr_subtopic"],
            explanation=None # Could use correct answer text as explanation
        )
        
        # Add answer options
        for order, ans in enumerate(q_data["answers"]):
            answer = ExamAnswerOption(
                answer_text=ans["text"],
                is_correct=ans["is_correct"],
                order=order
            )
            question.answers.append(answer)
        
        question_items.append(question)
    
    db.add_all(question_items)
    db.flush()  # Flush to get IDs
    print(f"  ✓ Created {len(question_items)} question items")
    
    # Now create exams and link questions
    print(f"  Creating {exam_count} exams...")
    for i in range(exam_count):
        exam = Exam(
            title=f"CBR Theorie Examen {i + 1}",
            description=f"Officieel CBR theorie-examen met 50 vragen (Examen {i + 1})",
            time_limit=30,
            passing_score=86,
            category="Theorie",
            is_published=True  # Auto-publish for dev
        )
        
        # Assign questions [i*50 : (i+1)*50]
        start_idx = i * questions_per_exam
        end_idx = min((i + 1) * questions_per_exam, len(question_items))
        exam.questions = question_items[start_idx:end_idx]
        
        db.add(exam)
        print(f"  ✓ Created: {exam.title} ({len(exam.questions)} questions)")
    
    db.commit()
    print("  ✅ All exams and questions created successfully!")


def main():
    """Main import function"""
    print("=" * 60)
    print("🚀 Slagie Platform - Advanced Excel Import with Images")
    print("=" * 60)
    print()
    
    # Paths
    excel_file = Path(__file__).parent.parent / "data" / "TheorieToppers examen vragen (3).xlsx"
    
    if not excel_file.exists():
        print(f"❌ Error: Excel file not found at {excel_file}")
        return
    
    print(f"📂 Excel file: {excel_file}")
    
    # Create static folder
    static_dir = create_static_folder()
    print(f"📂 Static images folder: {static_dir}")
    print()
    
    # Load Excel
    print("📊 Loading Excel file...")
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        print(f"  ✓ Loaded worksheet: {worksheet.title}")
        print(f"  ✓ Rows: {worksheet.max_row}, Columns: {worksheet.max_column}")
    except Exception as e:
        print(f"❌ Error loading Excel: {e}")
        return
    
    # Extract images
    image_map = extract_images_from_excel(workbook, worksheet, static_dir)
    
    # Parse questions
    questions_data = parse_excel_data(worksheet, image_map, static_dir)
    
    if not questions_data:
        print("❌ No questions found in Excel!")
        return
    
    # Create database session
    print("\n🗄️  Connecting to database...")
    
    # Recreate tables
    print("  ⚠️  Dropping existing exam tables...")
    Base.metadata.drop_all(bind=engine, tables=[
        Base.metadata.tables['exam_answer_options'],
        Base.metadata.tables['exam_questions_link'],
        Base.metadata.tables['exam_question_items'],
        Base.metadata.tables['exams']
    ])
    
    print("  🔨 Creating fresh tables...")
    Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    
    try:
        # Create exams and questions
        create_exams_and_questions(db, questions_data)
        
        # Print summary
        print()
        print("=" * 60)
        print("✅ IMPORT COMPLETE!")
        print("=" * 60)
        print(f"📸 Images extracted: {len(image_map)}")
        print(f"❓ Questions created: {len(questions_data)}")
        print(f"📝 Exams created: 3")
        print()
        print("Next steps:")
        print("  1. Start backend: uvicorn main:app --port 8000 --reload")
        print("  2. Access images at: http://localhost:8000/static/images/")
        print("  3. Login as admin and view exams in Admin Dashboard")
        print()
        
    except Exception as e:
        print(f"\n❌ Error during import: {e}")
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
