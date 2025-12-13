import openpyxl

wb = openpyxl.load_workbook('/app/data/TheorieToppers examen vragen (3).xlsx')
ws = wb.active

print("Header row (columns A-J):")
header = ws[1]
for i, cell in enumerate(header[:10], 1):
    col_letter = chr(64+i)
    print(f"  Col {col_letter}: {cell.value}")

print("\nRow 2 (first data row) - all columns:")
data_row = ws[2]
for i, cell in enumerate(data_row[:10], 1):
    col_letter = chr(64+i)
    val = str(cell.value)[:50] if cell.value else None
    print(f"  Col {col_letter}: {val}")

print(f"\nTotal rows in sheet: {ws.max_row}")
print(f"Total columns in sheet: {ws.max_column}")
